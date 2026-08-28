#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Construye index.html del Tablero Venta B2B — Café Quindío desde los Excel en data/.
Lo ejecuta GitHub Actions en cada cambio. Reproduce el pipeline B2B:
exclusión de vendedores no B2B, normalización de nombres, datos fila-a-fila
(vendedor, cliente, sucursal, línea, producto, tipología), lanzamiento y presupuesto.
"""
import openpyxl, json, re, unicodedata, datetime, glob, os, sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DATA = os.path.join(ROOT, "data")
CFG  = json.load(open(os.path.join(HERE, "config.json"), encoding="utf-8"))

def na(s):
    if s is None: return ""
    return unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode().upper().strip()

EXCL = {na(x) for x in CFG["excluidos"]} | {na("Genérico")}
RENAME = {"KEY ACCOUNT MANAGER": "KAM", "VENTAS CARTAGENA": "VENTAS COSTA"}
LAUNCH_REFS = CFG["launch_refs"]
REFN = {na(r): r for r in LAUNCH_REFS}

def kgf(item):
    u = (item or '').upper()
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*KG', u)
    if m: return float(m.group(1).replace(',', '.'))
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*G(?:R|RS)?\b', u)
    if m: return float(m.group(1).replace(',', '.')) / 1000.0
    return 0.0

def U(s): return unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode().upper()

def find(*keys, notkeys=()):
    for f in sorted(glob.glob(os.path.join(DATA, "*"))):
        u = U(os.path.basename(f))
        if not u.endswith(('.XLSX', '.XLS', '.XLSM')): continue
        if all(k in u for k in keys) and not any(nk in u for nk in notkeys):
            return f
    return None

def ci(hdr, name):
    for i, h in enumerate(hdr):
        if h and str(h).strip().lower() == name: return i
    return None

def tipidx(hdr):
    for i, h in enumerate(hdr):
        if h and ('tipolog' in str(h).lower() or 'tipo de cliente' in str(h).lower()): return i
    return None

def datekey(f):
    if isinstance(f, datetime.datetime): return f.strftime("%Y-%m-%d")
    s = str(f); m = re.search(r'(\d{4})-(\d{2})-(\d{2})', s)
    return m.group(0) if m else None

# ---------- localizar archivos ----------
AUG    = find('AGOSTO 2026', notkeys=('AVANCE', '2025'))
JUL    = find('JULIO 2026',  notkeys=('AVANCE', '2025'))
JUN    = find('JUNIO 2026',  notkeys=('AVANCE', '2025'))
AVANCE = find('AVANCE', 'AGOSTO')
Y2025  = find('2025', notkeys=('AVANCE',))
for nm, f in [("AGOSTO 2026", AUG), ("JULIO 2026", JUL), ("JUNIO 2026", JUN), ("AVANCE AGOSTO", AVANCE), ("INFORME 2025", Y2025)]:
    if not f:
        sys.exit("ERROR: no encontré el archivo para %s en data/. Revisa el nombre." % nm)
    print("  %-16s -> %s" % (nm, os.path.basename(f)))

# ---------- filas por mes (strings) ----------
allmonths = {}
def add_month_rows(path, sheet, only=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        sheet = "FAC. CONSOLIDADO" if "FAC. CONSOLIDADO" in wb.sheetnames else ("FAC CONSOLIDADA" if "FAC CONSOLIDADA" in wb.sheetnames else wb.sheetnames[0])
    ws = wb[sheet]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, max_col=16, values_only=True)))
    Fi = ci(hdr, 'fecha'); Vi = ci(hdr, 'nombre vendedor'); CLi = ci(hdr, 'razón social cliente factura')
    Si = ci(hdr, 'desc. sucursal despacho'); Li = ci(hdr, 'linea'); Ii = ci(hdr, 'desc. item')
    Qi = ci(hdr, 'cantidad inv.'); VALi = ci(hdr, 'valor subtotal'); Ti = tipidx(hdr)
    for r in ws.iter_rows(min_row=2, max_col=16, values_only=True):
        if Fi is None or r[Fi] is None: continue
        dk = datekey(r[Fi])
        if not dk: continue
        ym = dk[:7]
        if only and ym not in only: continue
        vv = str(r[Vi]).strip()
        if na(vv) in EXCL: continue
        vv = RENAME.get(vv.upper(), vv)
        try: val = float(r[VALi] or 0)
        except: val = 0.0
        try: q = float(r[Qi] or 0)
        except: q = 0.0
        day = int(dk[8:10])
        allmonths.setdefault(ym, []).append([
            vv,
            str(r[CLi]).strip() if r[CLi] else "SIN CLIENTE",
            str(r[Si]).strip() if (Si is not None and r[Si]) else "",
            str(r[Li]).strip() if r[Li] else "SIN LINEA",
            str(r[Ii]).strip() if r[Ii] else "SIN ITEM",
            str(r[Ti]).strip() if (Ti is not None and r[Ti]) else "SIN TIPO",
            day, round(val), round(q, 2)])
    wb.close()

print("Procesando ventas...")
add_month_rows(AUG, "FAC. CONSOLIDADO")
add_month_rows(JUL, "FAC. CONSOLIDADO")
add_month_rows(JUN, "FAC. CONSOLIDADO")
add_month_rows(Y2025, "FAC CONSOLIDADA", only={"2025-06", "2025-07", "2025-08"})

# ---------- indexar compacto ----------
V = {}; C = {}; S = {}; L = {}; I = {}; T = {}; itemkg = []
def gi(d, k):
    if k not in d: d[k] = len(d)
    return d[k]
months = {}
for ym, rows in allmonths.items():
    out = []
    for r in rows:
        vv, cl, su, li, it, tp, day, val, q = r
        ii = gi(I, it)
        if ii == len(itemkg): itemkg.append(round(kgf(it), 4))
        out.append([gi(V, vv), gi(C, cl), gi(S, su), gi(L, li), ii, gi(T, tp), day, val, q])
    months[ym] = out
def inv(d):
    a = [None] * len(d)
    for k, i in d.items(): a[i] = k
    return a
rowdata = {"dict": {"V": inv(V), "C": inv(C), "S": inv(S), "L": inv(L), "I": inv(I), "T": inv(T), "itemkg": itemkg}, "months": months}

# ---------- lanzamiento (Agosto 2026) ----------
wb = openpyxl.load_workbook(AUG, read_only=True, data_only=True); ws = wb["FAC. CONSOLIDADO"]
hdr = list(next(ws.iter_rows(min_row=1, max_row=1, max_col=16, values_only=True)))
F = ci(hdr, 'fecha'); Vi = ci(hdr, 'nombre vendedor'); Ii = ci(hdr, 'desc. item'); Qi = ci(hdr, 'cantidad inv.'); VALi = ci(hdr, 'valor subtotal'); CLi = ci(hdr, 'razón social cliente factura')
lrows = []
for r in ws.iter_rows(min_row=2, max_col=16, values_only=True):
    if r[F] is None: continue
    it = str(r[Ii]).strip() if r[Ii] else ""; ref = REFN.get(na(it))
    if not ref: continue
    vv = str(r[Vi]).strip()
    if na(vv) in EXCL: continue
    vv = RENAME.get(vv.upper(), vv)
    try: val = float(r[VALi] or 0)
    except: val = 0.0
    try: q = float(r[Qi] or 0)
    except: q = 0.0
    dk = datekey(r[F]) or ""
    lrows.append([vv, ref, str(r[CLi]).strip() if r[CLi] else "SIN CLIENTE", dk, round(val), q, round(kgf(it) * q, 2)])
wb.close()
launch = {"mes": "Agosto 2026", "refs": LAUNCH_REFS, "rows": lrows}

# ---------- presupuesto Agosto (Avance) ----------
from collections import defaultdict
wb = openpyxl.load_workbook(AVANCE, read_only=True, data_only=True)
ago = sorted((int(s.split()[1]), s) for s in wb.sheetnames if re.match(r'^AGO \d+$', s))
last = wb[ago[-1][1]]
bc = defaultdict(float); bv = defaultdict(float); bvc = defaultdict(lambda: defaultdict(float)); buf = []
for r in last.iter_rows(min_row=5, max_col=6, values_only=True):
    B = r[1]; A = r[0]; Fv = r[5]
    Bs = str(B).strip() if B is not None else ""; As = str(A).strip() if A is not None else ""
    if Bs == "Totales": buf = []; continue
    if Bs == "Totals":
        if As == "": buf = []; continue
        vv = RENAME.get(As.upper(), As)
        if na(As) not in EXCL:
            for ln, pr in buf: bc[ln] += pr; bv[vv] += pr; bvc[vv][ln] += pr
        buf = []; continue
    if " - " in Bs:
        try: pr = float(Fv or 0)
        except: pr = 0
        buf.append((Bs, pr))
wb.close()
budget_ago = {"bycat": dict(bc), "byvend": dict(bv), "byvendcat": {k: dict(v) for k, v in bvc.items()}, "total": sum(bc.values())}

# ---------- ensamblar BASE ----------
BASE = {
    "meta": {"generado": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
             "excluidos": CFG["excluidos"], "fuentes": CFG.get("fuentes", {})},
    "sel": CFG["sel"], "order": CFG["order"], "labels": CFG["labels"], "short": CFG["short"],
    "config": CFG["config"],
    "budgets": {"2026-08": budget_ago, "2026-07": CFG["budget_2026_07"]},
    "launch": launch,
    "rowdata": rowdata,
}

# ---------- inyectar en la plantilla ----------
tpl = open(os.path.join(ROOT, "web", "plantilla.html"), encoding="utf-8").read()
assert "/*__CQDATA__*/{}" in tpl, "La plantilla no tiene el marcador /*__CQDATA__*/{}"
outhtml = tpl.replace("/*__CQDATA__*/{}", json.dumps(BASE, ensure_ascii=False), 1)
open(os.path.join(ROOT, "index.html"), "w", encoding="utf-8").write(outhtml)

print("OK -> index.html  (%.2f MB)" % (len(outhtml) / 1048576))
print("Meses:", {k: len(v) for k, v in months.items()})
print("Agosto MTD:", round(sum(v[7] for v in months.get("2026-08", []))))
print("Presupuesto Agosto:", round(budget_ago["total"]), "| Lanzamiento filas:", len(lrows))
